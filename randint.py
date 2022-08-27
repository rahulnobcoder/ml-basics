import random

import openpyxl as op

x=op.load_workbook('rhul.xlsx')
sheet = x['Sheet1']
y=sheet.cell(1,1 )

for i in range(1,sheet.max_row +1):
    cell=sheet.cell(i,1)
    cell.value = int(random.randint(1,5))


x.save('rhul.xlsx')
